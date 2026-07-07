#!/usr/bin/env python3
"""
parse_excel.py — SWITCH Diagnostic Tool
Current Engineering

PURPOSE
───────
Converts a DTC Excel workbook into a JavaScript data file
that can be loaded by the web app.

Each Excel sheet = one Case.
Each row (after the header) = one diagnostic step.

EXCEL COLUMN FORMAT (required)
────────────────────────────────────
Col A  DTC code          e.g. 4
Col B  DTC name          e.g. HighVltgBattFault
Col C  DTC description   e.g. High Voltage Battery pack fault
Col D  DTC action        e.g. Shutdown the System
Col E  Areas of Issue    e.g. "Junction Box to Sub Pack Harness: Wire continuity check"
                              ← split on ":" → area_group / area_detail
Col F  Step              diagnostic instruction text
Col G  Observation       label for what the technician should record
Col H  Corrective Action what to do if the observation fails

INPUT TYPE AUTO-DETECTION (from Col G Observation text)
────────────────────────────────────────────────────────
"resistance", "value", "___"  →  "number"   (numeric text field, unit auto-detected)
"yes or no", "yes or n"       →  "yesno"    (YES / NO buttons)
"mounting", "condition",
"eyelet", "connector"         →  "condition" (OK / NOT OK buttons)
everything else               →  "yesno"    (default)

USAGE
─────
  # Single Excel file → single DTC:
  python parse_excel.py data/error_code_4.xlsx

  # Specify DTC number and output path explicitly:
  python parse_excel.py data/error_code_4.xlsx --dtc 4 --output data/dtc_4_cases.js

  # Dry run — print JSON only, don't write file:
  python parse_excel.py data/error_code_4.xlsx --dry-run

OUTPUT
──────
Creates  data/dtc_<N>_cases.js  containing:
  const DTC_<N>_CASES = [ ...array of case objects... ];

Then in index.html add:
  <script src="../data/dtc_<N>_cases.js"></script>

And in app.js under DTC_LIBRARY add:
  "<N>": {
    code: "<N>", name: "...", description: "...",
    action: "...", severity: "CRITICAL", category: "...",
    vehicles: ["BUS","LCV"],
    cases: DTC_<N>_CASES
  }
"""

import sys
import os
import json
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Install it with:  pip install openpyxl")
    sys.exit(1)


def clean(value):
    """Strip and stringify a cell value, return empty string for None."""
    if value is None:
        return ''
    return str(value).strip()


def detect_input_type(obs_label: str) -> str:
    """
    Determine the UI input widget from the Observation column text.

    Returns:
        "yesno"     → YES / NO toggle buttons
        "condition" → OK / NOT OK toggle buttons
        "number"    → numeric text field (for resistance, voltage, etc.)
    """
    lbl = obs_label.lower()

    if 'resistance' in lbl or 'value' in lbl or '___' in obs_label:
        return 'number'

    if 'yes or no' in lbl or 'yes or n' in lbl:
        return 'yesno'

    if any(kw in lbl for kw in ['mounting', 'condition', 'eyelet', 'connector']):
        return 'condition'

    return 'yesno'  # default


def parse_workbook(path: str) -> list:
    """
    Parse an Excel workbook and return a list of case objects.

    Args:
        path: Path to the .xlsx file

    Returns:
        List of case dicts, one per sheet.
    """
    wb = load_workbook(path, read_only=True)
    cases = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(max_row=60, values_only=True))

        if len(all_rows) < 2:
            print(f"  ⚠  Skipping '{sheet_name}' — fewer than 2 rows")
            continue

        steps      = []
        dtc_code   = None
        dtc_name   = None
        dtc_desc   = None
        dtc_action = None
        area       = None

        # Skip header row (index 0), process data rows
        for row in all_rows[1:]:
            # Pad row to at least 8 columns
            cells = [clean(row[i]) if i < len(row) else '' for i in range(8)]

            # Carry forward first-seen values (cells may be empty due to Excel merging)
            if cells[0] and not dtc_code:   dtc_code   = cells[0]
            if cells[1] and not dtc_name:   dtc_name   = cells[1]
            if cells[2] and not dtc_desc:   dtc_desc   = cells[2]
            if cells[3] and not dtc_action: dtc_action = cells[3]
            if cells[4]:                    area       = cells[4]

            step_text  = cells[5]
            obs_label  = cells[6]
            corrective = cells[7]

            # Skip empty or NA rows
            if not step_text or step_text.upper() == 'NA':
                continue

            # Tidy up the observation label
            obs_clean = obs_label.strip().rstrip(':').strip()

            steps.append({
                'step':        step_text,
                'obs_label':   obs_clean,
                'corrective':  corrective,
                'input_type':  detect_input_type(obs_label)
            })

        if not steps:
            print(f"  ⚠  Skipping '{sheet_name}' — no valid steps found")
            continue

        # Split area into group and detail on ':'
        area_group  = area.split(':')[0].strip()  if area and ':' in area else (area or sheet_name)
        area_detail = area.split(':')[-1].strip() if area and ':' in area else ''

        case = {
            'sheet':       sheet_name,
            'area_group':  area_group,
            'area_detail': area_detail,
            'steps':       steps
        }

        # Store top-level DTC info on first case (used for reference)
        if dtc_code:
            case['_dtc_code']   = dtc_code
            case['_dtc_name']   = dtc_name
            case['_dtc_desc']   = dtc_desc
            case['_dtc_action'] = dtc_action

        cases.append(case)
        print(f"  ✓  {sheet_name}: {len(steps)} steps — {area_group} / {area_detail}")

    return cases


def main():
    parser = argparse.ArgumentParser(
        description='Parse DTC Excel workbook into JavaScript data file for SWITCH Diagnostic Tool'
    )
    parser.add_argument('excel',           help='Path to input .xlsx file')
    parser.add_argument('--dtc',           help='DTC number (e.g. 4). Auto-detected from file if omitted.')
    parser.add_argument('--output', '-o',  help='Output .js file path. Defaults to data/dtc_<N>_cases.js')
    parser.add_argument('--dry-run',       action='store_true', help='Print JSON only, do not write file')
    args = parser.parse_args()

    # ── Validate input ──
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    print(f"\nParsing: {excel_path.name}")
    print("─" * 50)

    # ── Parse ──
    cases = parse_workbook(str(excel_path))

    if not cases:
        print("\nERROR: No cases could be parsed. Check your Excel file format.")
        sys.exit(1)

    # ── Determine DTC number ──
    dtc_num = args.dtc
    if not dtc_num:
        # Try to get from first case
        first_dtc = cases[0].get('_dtc_code', '')
        if first_dtc:
            dtc_num = str(first_dtc).split('.')[0]  # Remove decimal if any
        else:
            dtc_num = excel_path.stem.replace('error_code_', '').replace('dtc_', '')

    # Clean the DTC number for use as JS variable name
    dtc_var = dtc_num.replace('-', '_').replace('.', '_')

    print(f"\n{'─' * 50}")
    print(f"DTC:    {dtc_num}")
    print(f"Cases:  {len(cases)}")
    total_steps = sum(len(c['steps']) for c in cases)
    print(f"Steps:  {total_steps} total")

    # ── Strip internal _dtc_* fields from output ──
    clean_cases = []
    for c in cases:
        clean = {k: v for k, v in c.items() if not k.startswith('_')}
        clean_cases.append(clean)

    # ── Generate JS ──
    cases_json = json.dumps(clean_cases, indent=2, ensure_ascii=False)

    js_content = f"""/* ═══════════════════════════════════════════════════════
   DTC {dtc_num} Cases — Auto-generated by scripts/parse_excel.py
   Source: {excel_path.name}
   Cases: {len(cases)}
   Steps: {total_steps}

   DO NOT EDIT MANUALLY.
   Re-generate by running:
     python scripts/parse_excel.py data/{excel_path.name}
═══════════════════════════════════════════════════════ */

const DTC_{dtc_var}_CASES = {cases_json};
"""

    if args.dry_run:
        print("\nDRY RUN — output below:\n")
        print(js_content[:2000])
        if len(js_content) > 2000:
            print(f"... [{len(js_content):,} chars total]")
        return

    # ── Write output ──
    if args.output:
        out_path = Path(args.output)
    else:
        # Default: next to the Excel file in data/
        out_path = excel_path.parent / f"dtc_{dtc_var}_cases.js"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(js_content, encoding='utf-8')

    print(f"\n✓ Written: {out_path}  ({len(js_content):,} chars)")
    print(f"\nNEXT STEPS:")
    print(f"  1. In app/index.html, add:")
    print(f"       <script src=\"../data/dtc_{dtc_var}_cases.js\"></script>")
    print(f"  2. In app/app.js under DTC_LIBRARY, add:")
    print(f'       "{dtc_num}": {{')
    print(f'         code: "{dtc_num}", name: "FaultName",')
    print(f'         description: "Full description",')
    print(f'         action: "Shutdown the System",')
    print(f'         severity: "CRITICAL",')
    print(f'         category: "HV Battery",')
    print(f'         vehicles: ["BUS","LCV"],')
    print(f'         cases: DTC_{dtc_var}_CASES')
    print(f'       }}')


if __name__ == '__main__':
    main()
